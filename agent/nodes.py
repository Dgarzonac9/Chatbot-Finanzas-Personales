import io
import psycopg2, json, os
import psycopg2.extras
import psycopg2.pool
from contextlib import contextmanager
from datetime import date
from dotenv import load_dotenv
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL")
MODELO_LLM   = os.getenv("MODELO_LLM")
API_KEY      = os.getenv("GROQ_API_KEY")

client = Groq(api_key=API_KEY)

# ── Connection pool (lazy) ─────────────────────────────────────────────────

_pool = None

def get_pool():
    global _pool
    if _pool is None or _pool.closed:
        _pool = psycopg2.pool.ThreadedConnectionPool(
            minconn=1,
            maxconn=10,
            dsn=DATABASE_URL,
            cursor_factory=psycopg2.extras.DictCursor,
        )
    return _pool

@contextmanager
def get_conn():
    conn = get_pool().getconn()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        get_pool().putconn(conn)

# ── Helpers internos ───────────────────────────────────────────────────────

def _buscar_vaca(cur, user_id: int, nombre_vaca: str):
    cur.execute(
        "SELECT id, nombre, num_personas FROM vacas WHERE user_id = %s AND LOWER(nombre) LIKE LOWER(%s) ORDER BY id DESC LIMIT 1",
        (user_id, f"%{nombre_vaca}%"),
    )
    row = cur.fetchone()
    if row:
        return row
    palabras = [p for p in nombre_vaca.split() if len(p) > 3]
    for palabra in palabras:
        cur.execute(
            "SELECT id, nombre, num_personas FROM vacas WHERE user_id = %s AND LOWER(nombre) LIKE LOWER(%s) ORDER BY id DESC LIMIT 1",
            (user_id, f"%{palabra}%"),
        )
        row = cur.fetchone()
        if row:
            return row
    return None

def _listar_vacas(cur, user_id: int) -> str:
    cur.execute(
        "SELECT id, nombre FROM vacas WHERE user_id = %s AND cerrada = FALSE ORDER BY id DESC",
        (user_id,),
    )
    vacas = cur.fetchall()
    if not vacas:
        return None
    return "\n".join([f"  {i+1}️⃣ *{v[1]}* (ID: {v[0]})" for i, v in enumerate(vacas)])

def _validar_user_id(state: dict) -> int:
    user_id = state.get("user_id")
    if not isinstance(user_id, int) or user_id <= 0:
        raise ValueError(f"user_id inválido: {user_id!r}")
    return user_id

def formatear_respuesta(datos_crudos: str) -> str:
    prompt = f"""
    Eres un asistente de finanzas personales amigable.
    Convierte estos datos en un mensaje claro y amigable en español.
    Datos: {datos_crudos}
    Responde de forma breve y directa, puedes usar emojis.
    """
    resp = client.chat.completions.create(
        model=MODELO_LLM,
        messages=[{"role": "user", "content": prompt}],
    )
    return resp.choices[0].message.content.strip()


# ── Nodo 1: Guardar gasto (soporta múltiples gastos) ──────────────────────

def extraer_gastos_llm(texto: str):
    prompt = f"""
    El usuario puede escribir con errores ortográficos o de forma informal.
    Puede mencionar UNO o VARIOS gastos en el mismo mensaje.

    Extrae TODOS los gastos en JSON:
    {{
      "gastos": [
        {{"categoria": "...", "monto": ..., "fecha": "YYYY-MM-DD"}},
        ...
      ]
    }}
    Categorías posibles: comida, transporte, entretenimiento, salud, ropa, servicios, otro.
    Si no menciona fecha usa hoy: {date.today()}
    Si escribe "10k" interpreta como 10000, "20k" como 20000, etc.
    IMPORTANTE: extrae SOLO gastos explícitos con descripción y monto numérico.
    Texto: "{texto}"
    """
    resp = client.chat.completions.create(
        model=MODELO_LLM,
        messages=[{"role": "user", "content": prompt}],
        response_format={"type": "json_object"},
    )
    try:
        data = json.loads(resp.choices[0].message.content)
        gastos = data.get("gastos", [])
        if not gastos and data.get("categoria"):
            gastos = [data]
        return gastos
    except Exception as e:
        print("Error extrayendo gastos:", e)
        return []

def guardar_gasto(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        gastos = extraer_gastos_llm(state["input"])

        if not gastos:
            return {"output": "No pude entender el gasto. Intenta: 'gasté 10k en comida'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                for g in gastos:
                    cur.execute(
                        "INSERT INTO gastos (user_id, categoria, monto, fecha) VALUES (%s, %s, %s, %s)",
                        (user_id, g.get("categoria", "otro"), float(g.get("monto", 0)), g.get("fecha", str(date.today()))),
                    )

        if len(gastos) == 1:
            g = gastos[0]
            respuesta = formatear_respuesta(f"Gasto guardado: {g['monto']} COP en '{g['categoria']}' el {g['fecha']}")
        else:
            desglose = ", ".join([f"{g['categoria']}: ${int(float(g['monto'])):,}" for g in gastos])
            total = sum(float(g["monto"]) for g in gastos)
            respuesta = formatear_respuesta(f"Guardé {len(gastos)} gastos: {desglose}. Total: ${int(total):,}")

        return {"output": respuesta}
    except ValueError as e:
        return {"output": f"⚠️ {str(e)}"}
    except psycopg2.Error as e:
        print(f"[DB Error guardar_gasto] {e}")
        return {"output": "❌ Hubo un problema con la base de datos. Intenta en un momento."}
    except Exception as e:
        print(f"[Error guardar_gasto] {e}")
        return {"output": "❌ Error guardando el gasto. Intenta de nuevo."}


# ── Nodo 2: Reporte del día ────────────────────────────────────────────────

def reporte_dia(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        hoy = str(date.today())

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT categoria, SUM(monto) AS total FROM gastos WHERE user_id = %s AND fecha = %s GROUP BY categoria ORDER BY total DESC",
                    (user_id, hoy),
                )
                filas = cur.fetchall()
                cur.execute("SELECT COALESCE(SUM(monto), 0) FROM gastos WHERE user_id = %s AND fecha = %s", (user_id, hoy))
                total = cur.fetchone()[0]

        if not filas:
            return {"output": "No registré gastos tuyos hoy 🎉"}

        desglose = ", ".join([f"{cat}: ${round(m):,}" for cat, m in filas])
        return {"output": formatear_respuesta(f"Reporte del {hoy}. Total: ${round(total):,}. Desglose: {desglose}")}
    except Exception as e:
        print(f"[Error reporte_dia] {e}")
        return {"output": "❌ Error generando el reporte del día."}


# ── Nodo 3: Reporte del mes ────────────────────────────────────────────────

def reporte_mes(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        mes = date.today().strftime("%Y-%m")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT categoria, SUM(monto) AS total FROM gastos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s GROUP BY categoria ORDER BY total DESC", (user_id, mes))
                filas = cur.fetchall()
                cur.execute("SELECT COALESCE(SUM(monto), 0) FROM gastos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s", (user_id, mes))
                total = cur.fetchone()[0]
                cur.execute("SELECT fecha, SUM(monto) AS t FROM gastos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s GROUP BY fecha ORDER BY t DESC LIMIT 1", (user_id, mes))
                dia_top = cur.fetchone()

        if not filas:
            return {"output": f"Sin gastos registrados en {mes} 🎉"}

        desglose = ", ".join([f"{cat}: ${round(m):,}" for cat, m in filas])
        extra = f"Día con más gasto: {dia_top[0]} (${round(dia_top[1]):,})" if dia_top else ""
        return {"output": formatear_respuesta(f"Reporte {mes}. Total: ${round(total):,}. {extra}. Por categoría: {desglose}")}
    except Exception as e:
        print(f"[Error reporte_mes] {e}")
        return {"output": "❌ Error generando el reporte del mes."}


# ── Nodo 4: Reporte por categoría ─────────────────────────────────────────

def reporte_categoria(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT categoria, SUM(monto) AS total, COUNT(*) AS veces FROM gastos WHERE user_id = %s GROUP BY categoria ORDER BY total DESC", (user_id,))
                filas = cur.fetchall()

        if not filas:
            return {"output": "Aún no tienes gastos registrados 📊"}

        resumen = "; ".join([f"{cat}: ${round(t):,} ({v} veces)" for cat, t, v in filas])
        return {"output": formatear_respuesta(f"Historial de gastos por categoría: {resumen}")}
    except Exception as e:
        print(f"[Error reporte_categoria] {e}")
        return {"output": "❌ Error generando el reporte por categoría."}


# ── Nodo 5: Editar / eliminar gasto ───────────────────────────────────────

def editar_gasto(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere eliminar un gasto. Extrae en JSON:
        {{"accion": "ultimo" | "por_categoria" | "por_fecha",
          "categoria": "..." (opcional),
          "fecha": "YYYY-MM-DD" (opcional)}}
        Texto: "{texto}"
        Fecha de hoy: {date.today()}
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
        except Exception:
            params = {"accion": "ultimo"}

        accion = params.get("accion", "ultimo")

        with get_conn() as conn:
            with conn.cursor() as cur:
                if accion == "ultimo":
                    cur.execute("SELECT id FROM gastos WHERE user_id = %s ORDER BY id DESC LIMIT 1", (user_id,))
                    row = cur.fetchone()
                    if row:
                        cur.execute("DELETE FROM gastos WHERE id = %s", (row["id"],))
                        return {"output": "✅ Eliminé tu último gasto registrado."}
                elif accion == "por_categoria" and params.get("categoria"):
                    fecha = params.get("fecha", str(date.today()))
                    cur.execute("DELETE FROM gastos WHERE user_id = %s AND categoria = %s AND fecha = %s", (user_id, params["categoria"], fecha))
                    if cur.rowcount:
                        return {"output": f"✅ Eliminé los gastos de '{params['categoria']}' del {fecha}."}
                elif accion == "por_fecha" and params.get("fecha"):
                    cur.execute("DELETE FROM gastos WHERE user_id = %s AND fecha = %s", (user_id, params["fecha"]))
                    if cur.rowcount:
                        return {"output": f"✅ Eliminé los gastos del día {params['fecha']}."}

        return {"output": "No encontré el gasto que mencionas. ¿Puedes ser más específico?"}
    except Exception as e:
        print(f"[Error editar_gasto] {e}")
        return {"output": "❌ Error eliminando el gasto."}


# ── Nodo 6: Presupuesto ────────────────────────────────────────────────────

def presupuesto(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]
        mes = date.today().strftime("%Y-%m")

        prompt = f"""
        El usuario quiere establecer o consultar un presupuesto mensual.
        Extrae en JSON: {{"accion": "establecer" | "consultar", "monto": ... (solo si establece)}}
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
        except Exception:
            params = {"accion": "consultar"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                if params.get("accion") == "establecer" and params.get("monto"):
                    cur.execute(
                        "INSERT INTO presupuestos (user_id, mes, monto) VALUES (%s, %s, %s) ON CONFLICT (user_id, mes) DO UPDATE SET monto = EXCLUDED.monto",
                        (user_id, mes, float(params["monto"])),
                    )
                    presupuesto_val = float(params["monto"])
                else:
                    cur.execute("SELECT monto FROM presupuestos WHERE user_id = %s AND mes = %s", (user_id, mes))
                    row = cur.fetchone()
                    if not row:
                        return {"output": f"No tienes presupuesto definido para {mes}. Dime cuánto quieres gastar este mes y lo guardo 💰"}
                    presupuesto_val = float(row["monto"])

                cur.execute("SELECT COALESCE(SUM(monto), 0) FROM gastos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s", (user_id, mes))
                gastado = float(cur.fetchone()[0])

        porcentaje = (gastado / presupuesto_val * 100) if presupuesto_val else 0
        restante = presupuesto_val - gastado
        alerta = (
            "⚠️ ¡Superaste el presupuesto!"       if gastado > presupuesto_val else
            "🟡 Vas por más del 80%, con cuidado." if porcentaje > 80          else
            "✅ Vas bien."
        )
        return {"output": formatear_respuesta(
            f"Presupuesto {mes}: ${round(presupuesto_val):,}. Gastado: ${round(gastado):,} ({round(porcentaje)}%). Restante: ${round(restante):,}. {alerta}"
        )}
    except Exception as e:
        print(f"[Error presupuesto] {e}")
        return {"output": "❌ Error consultando el presupuesto."}


# ── Nodo 7: Crear vaca ─────────────────────────────────────────────────────

def crear_vaca(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere crear una vaca o fondo grupal.
        Extrae en JSON: {{"nombre": "..."}}
        El nombre es SOLO el nombre del fondo, sin palabras como "vaca", "crear", "fondo", "nuevo".
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            nombre = params.get("nombre", "Sin nombre")
        except Exception:
            nombre = "Sin nombre"

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("INSERT INTO vacas (user_id, nombre) VALUES (%s, %s) RETURNING id", (user_id, nombre))
                vaca_id = cur.fetchone()[0]

        return {"output": f"✅ Vaca *{nombre}* creada (ID: {vaca_id}).\nAgrégale gastos así:\n'agregar a vaca {nombre}: comida 50000, sonido 20000'"}
    except Exception as e:
        print(f"[Error crear_vaca] {e}")
        return {"output": "❌ Error creando la vaca."}


# ── Nodo 8: Agregar gastos a vaca ─────────────────────────────────────────

def agregar_vaca(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere agregar gastos a una vaca (fondo grupal).
        El formato puede ser: 'agregar a vaca NOMBRE: item1 monto1, item2 monto2'

        REGLAS IMPORTANTES:
        - Extrae SOLO gastos con descripción clara Y monto numérico explícito.
        - NO interpretes números de personas como gastos (ej: "12 personas" → no es gasto).
        - Si un número no tiene descripción de item asociada, ignóralo.
        - Solo extrae pares (descripcion, monto) claramente identificables.

        Extrae en JSON:
        {{
          "nombre_vaca": "...",
          "gastos": [{{"descripcion": "...", "monto": ...}}, ...]
        }}
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            nombre_vaca = params.get("nombre_vaca", "")
            gastos = params.get("gastos", [])
        except Exception:
            return {"output": "No pude entender. Intenta: 'agregar a vaca Salida: comida 50000, transporte 20000'"}

        if not gastos:
            return {"output": "No encontré gastos claros. Intenta: 'agregar a vaca Salida: comida 50000'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                row = _buscar_vaca(cur, user_id, nombre_vaca)
                if not row:
                    lista = _listar_vacas(cur, user_id)
                    if lista:
                        return {"output": f"No encontré esa vaca. Tus vacas activas son:\n{lista}"}
                    return {"output": "No tienes vacas activas. Crea una con 'crear vaca Nombre'"}

                vaca_id = row[0]
                nombre_real = row[1]

                for g in gastos:
                    cur.execute(
                        "INSERT INTO vaca_gastos (vaca_id, descripcion, monto) VALUES (%s, %s, %s)",
                        (vaca_id, g.get("descripcion", "gasto"), float(g.get("monto", 0))),
                    )

                cur.execute("SELECT COALESCE(SUM(monto),0) FROM vaca_gastos WHERE vaca_id = %s", (vaca_id,))
                total = cur.fetchone()[0]

        desglose = ", ".join([f"{g['descripcion']}: ${int(float(g['monto'])):,}" for g in gastos])
        return {"output": f"✅ Gastos agregados a *{nombre_real}*:\n{desglose}\n\n💰 Total acumulado: ${int(total):,}"}
    except Exception as e:
        print(f"[Error agregar_vaca] {e}")
        return {"output": "❌ Error agregando gastos a la vaca."}


# ── Nodo 9: Dividir vaca ───────────────────────────────────────────────────

def dividir_vaca(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere dividir una vaca entre personas.
        Extrae en JSON: {{"nombre_vaca": "...", "num_personas": ...}}
        num_personas debe ser un número entero mayor a 0.
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            nombre_vaca = params.get("nombre_vaca", "")
            num_personas = max(1, int(params.get("num_personas", 1)))
        except Exception:
            return {"output": "No pude entender. Intenta: 'dividir vaca Salida entre 5 personas'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, nombre FROM vacas WHERE user_id = %s AND LOWER(nombre) LIKE LOWER(%s) ORDER BY id DESC LIMIT 1",
                    (user_id, f"%{nombre_vaca}%"),
                )
                row = cur.fetchone()
                if not row:
                    return {"output": f"No encontré la vaca '{nombre_vaca}'."}
                vaca_id, nombre_real = row[0], row[1]

                cur.execute("SELECT descripcion, monto FROM vaca_gastos WHERE vaca_id = %s ORDER BY id", (vaca_id,))
                gastos = cur.fetchall()
                total = sum(float(g[1]) for g in gastos)
                cur.execute("UPDATE vacas SET num_personas = %s WHERE id = %s", (num_personas, vaca_id))

        if not gastos:
            return {"output": f"La vaca '{nombre_real}' no tiene gastos registrados."}

        por_persona = total / num_personas
        desglose = "\n".join([f"  • {g[0]}: ${int(float(g[1])):,}" for g in gastos])

        return {"output": (
            f"📊 *Vaca: {nombre_real}*\n\n"
            f"{desglose}\n\n"
            f"💰 Total: ${int(total):,}\n"
            f"👥 Personas: {num_personas}\n"
            f"➗ Cada uno pone: *${int(por_persona):,}*"
        )}
    except Exception as e:
        print(f"[Error dividir_vaca] {e}")
        return {"output": "❌ Error dividiendo la vaca."}


# ── Nodo 10: Resumen vaca ──────────────────────────────────────────────────

def resumen_vaca(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere ver el resumen de una vaca.
        Extrae en JSON: {{"nombre_vaca": "..."}}
        Si no menciona nombre específico usa: {{"nombre_vaca": "ultima"}}
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            nombre_vaca = params.get("nombre_vaca", "ultima")
        except Exception:
            nombre_vaca = "ultima"

        with get_conn() as conn:
            with conn.cursor() as cur:
                if nombre_vaca == "ultima":
                    cur.execute("SELECT id, nombre, num_personas FROM vacas WHERE user_id = %s ORDER BY id DESC LIMIT 1", (user_id,))
                else:
                    cur.execute(
                        "SELECT id, nombre, num_personas FROM vacas WHERE user_id = %s AND LOWER(nombre) LIKE LOWER(%s) ORDER BY id DESC LIMIT 1",
                        (user_id, f"%{nombre_vaca}%"),
                    )
                row = cur.fetchone()
                if not row:
                    return {"output": "No encontré ninguna vaca. Crea una con 'crear vaca Nombre'"}
                vaca_id, nombre_real, num_personas = row

                cur.execute("SELECT descripcion, monto FROM vaca_gastos WHERE vaca_id = %s ORDER BY id", (vaca_id,))
                gastos = cur.fetchall()
                total = sum(float(g[1]) for g in gastos)

        if not gastos:
            return {"output": f"La vaca *{nombre_real}* no tiene gastos aún."}

        desglose = "\n".join([f"  • {g[0]}: ${int(float(g[1])):,}" for g in gastos])
        por_persona = f"${int(total/num_personas):,}" if num_personas > 1 else "aún no dividida"

        return {"output": (
            f"📊 *Resumen: {nombre_real}*\n\n"
            f"{desglose}\n\n"
            f"💰 Total: ${int(total):,}\n"
            f"👥 Personas: {num_personas}\n"
            f"➗ Por persona: {por_persona}"
        )}
    except Exception as e:
        print(f"[Error resumen_vaca] {e}")
        return {"output": "❌ Error obteniendo el resumen de la vaca."}


# ── Nodo 11: Mis Vacas ─────────────────────────────────────────────────────

def mis_vacas(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT v.id, v.nombre, v.num_personas,
                           COALESCE(SUM(vg.monto), 0) as total
                    FROM vacas v
                    LEFT JOIN vaca_gastos vg ON v.id = vg.id
                    WHERE v.user_id = %s AND v.cerrada = FALSE
                    GROUP BY v.id, v.nombre, v.num_personas
                    ORDER BY v.id DESC
                    """,
                    (user_id,),
                )
                vacas = cur.fetchall()

        if not vacas:
            return {"output": "No tienes vacas activas. Crea una con 'crear vaca Nombre' 🐄"}

        lista = "\n".join([
            f"  {i+1}️⃣ *{v[1]}* — Total: ${int(v[3]):,} | 👥 {v[2]} personas"
            for i, v in enumerate(vacas)
        ])
        return {"output": f"🐄 *Tus vacas activas:*\n\n{lista}\n\nPara agregar gastos:\n'agregar a vaca [nombre]: item monto'"}
    except Exception as e:
        print(f"[Error mis_vacas] {e}")
        return {"output": "❌ Error listando las vacas."}


# ── Nodo 12: Registrar deuda ───────────────────────────────────────────────

def registrar_deuda(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere registrar una deuda informal (sin seguimiento de abonos).
        Extrae en JSON:
        {{
          "tipo": "prestado" (yo le presté a alguien) | "debo" (yo le debo a alguien),
          "persona": "...",
          "monto": ...,
          "descripcion": "..."
        }}
        Si escribe "10k" interpreta como 10000.
        Texto: "{texto}"
        Fecha hoy: {date.today()}
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
        except Exception:
            return {"output": "No pude entender. Intenta: 'le debo 50000 a Juan por el bus'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO deudas (user_id, tipo, persona, monto, descripcion) VALUES (%s, %s, %s, %s, %s)",
                    (user_id, params.get("tipo"), params.get("persona"), float(params.get("monto", 0)), params.get("descripcion", "")),
                )

        tipo    = params.get("tipo")
        persona = params.get("persona")
        monto   = int(float(params.get("monto", 0)))
        desc    = params.get("descripcion", "")

        if tipo == "prestado":
            return {"output": f"✅ Registrado: *{persona}* te debe ${monto:,}\n📝 {desc}"}
        else:
            return {"output": f"✅ Registrado: le debes ${monto:,} a *{persona}*\n📝 {desc}"}
    except Exception as e:
        print(f"[Error registrar_deuda] {e}")
        return {"output": "❌ Error registrando la deuda."}


# ── Nodo 13: Consultar deudas ──────────────────────────────────────────────

def consultar_deudas(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT tipo, persona, monto, descripcion, fecha FROM deudas WHERE user_id = %s AND pagado = FALSE ORDER BY fecha DESC",
                    (user_id,),
                )
                deudas = cur.fetchall()

        if not deudas:
            return {"output": "No tienes deudas pendientes 🎉"}

        prestados = [(p, m, d, f) for t, p, m, d, f in deudas if t == "prestado"]
        debo      = [(p, m, d, f) for t, p, m, d, f in deudas if t == "debo"]

        respuesta = "📋 *Deudas pendientes*\n\n"
        if prestados:
            total = sum(float(m) for _, m, _, _ in prestados)
            respuesta += f"💚 *Te deben (total: ${int(total):,})*\n"
            for p, m, d, f in prestados:
                respuesta += f"  • {p}: ${int(float(m)):,} — {d}\n"
        if debo:
            total = sum(float(m) for _, m, _, _ in debo)
            respuesta += f"\n🔴 *Debes (total: ${int(total):,})*\n"
            for p, m, d, f in debo:
                respuesta += f"  • {p}: ${int(float(m)):,} — {d}\n"

        return {"output": respuesta}
    except Exception as e:
        print(f"[Error consultar_deudas] {e}")
        return {"output": "❌ Error consultando las deudas."}


# ── Nodo 14: Pagar deuda ───────────────────────────────────────────────────

def pagar_deuda(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere marcar una deuda como pagada.
        Extrae en JSON: {{"persona": "..."}}
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            persona = params.get("persona", "")
        except Exception:
            return {"output": "No pude entender. Intenta: 'Juan me pagó' o 'pagué a María'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE deudas SET pagado = TRUE WHERE user_id = %s AND LOWER(persona) LIKE LOWER(%s) AND pagado = FALSE",
                    (user_id, f"%{persona}%"),
                )
                filas = cur.rowcount

        if filas:
            return {"output": f"✅ Deuda con *{persona}* marcada como pagada 🎉"}
        return {"output": f"No encontré deuda pendiente con '{persona}'."}
    except Exception as e:
        print(f"[Error pagar_deuda] {e}")
        return {"output": "❌ Error marcando la deuda como pagada."}


# ── Nodo 15: Registrar ingreso ─────────────────────────────────────────────

def registrar_ingreso(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere registrar un ingreso de dinero.
        Extrae en JSON:
        {{
          "descripcion": "...",
          "monto": ...,
          "fecha": "YYYY-MM-DD"
        }}
        Si no hay fecha usa hoy: {date.today()}
        Si escribe "10k" interpreta como 10000.
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            descripcion = params.get("descripcion", "ingreso")
            monto = float(params.get("monto", 0))
            fecha = params.get("fecha", str(date.today()))
        except Exception:
            return {"output": "No pude entender. Intenta: 'recibí salario 2500000'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO ingresos (user_id, descripcion, monto, fecha) VALUES (%s, %s, %s, %s)",
                    (user_id, descripcion, monto, fecha),
                )

        return {"output": f"✅ Ingreso registrado\n💵 {descripcion}: ${int(monto):,}\n📅 Fecha: {fecha}"}
    except Exception as e:
        print(f"[Error registrar_ingreso] {e}")
        return {"output": "❌ Error registrando el ingreso."}


# ── Nodo 16: Ver ingresos del mes ──────────────────────────────────────────

def ver_ingresos(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        mes = date.today().strftime("%Y-%m")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT descripcion, monto, fecha FROM ingresos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s ORDER BY fecha DESC",
                    (user_id, mes),
                )
                ingresos = cur.fetchall()
                total = sum(float(i[1]) for i in ingresos)

        if not ingresos:
            return {"output": f"No registraste ingresos en {mes} 📭"}

        detalle = "\n".join([f"  • {i[0]}: ${int(float(i[1])):,} ({i[2]})" for i in ingresos])
        return {"output": f"💵 *Ingresos de {mes}*\n\n{detalle}\n\n💰 Total: ${int(total):,}"}
    except Exception as e:
        print(f"[Error ver_ingresos] {e}")
        return {"output": "❌ Error consultando los ingresos."}


# ── Nodo 17: Balance del mes ───────────────────────────────────────────────

def balance_mes(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        mes = date.today().strftime("%Y-%m")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COALESCE(SUM(monto), 0) FROM ingresos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s", (user_id, mes))
                total_ingresos = float(cur.fetchone()[0])
                cur.execute("SELECT COALESCE(SUM(monto), 0) FROM gastos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s", (user_id, mes))
                total_gastos = float(cur.fetchone()[0])

        balance = total_ingresos - total_gastos
        emoji = "✅" if balance >= 0 else "🔴"
        estado = "positivo" if balance >= 0 else "negativo"

        return {"output": (
            f"📊 *Balance {mes}*\n\n"
            f"💵 Ingresos:  ${int(total_ingresos):,}\n"
            f"💸 Gastos:    ${int(total_gastos):,}\n"
            f"─────────────────\n"
            f"{emoji} Balance {estado}: ${int(balance):,}"
        )}
    except Exception as e:
        print(f"[Error balance_mes] {e}")
        return {"output": "❌ Error calculando el balance."}


# ── Nodo 18: Generar Excel ─────────────────────────────────────────────────

def generar_excel(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere un reporte Excel de sus finanzas.
        Extrae en JSON: {{"mes": "YYYY-MM"}}
        Si no menciona mes específico usa el mes actual: {date.today().strftime("%Y-%m")}
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            mes = params.get("mes", date.today().strftime("%Y-%m"))
        except Exception:
            mes = date.today().strftime("%Y-%m")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT categoria, monto, fecha FROM gastos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s ORDER BY fecha DESC", (user_id, mes))
                gastos = cur.fetchall()
                cur.execute("SELECT descripcion, monto, fecha FROM ingresos WHERE user_id = %s AND TO_CHAR(fecha, 'YYYY-MM') = %s ORDER BY fecha DESC", (user_id, mes))
                ingresos = cur.fetchall()
                cur.execute("SELECT tipo, persona, monto, descripcion, fecha FROM deudas WHERE user_id = %s AND pagado = FALSE ORDER BY fecha DESC", (user_id,))
                deudas = cur.fetchall()
                cur.execute(
                    "SELECT v.nombre, v.num_personas, COALESCE(SUM(vg.monto), 0) FROM vacas v LEFT JOIN vaca_gastos vg ON v.id = vg.id WHERE v.user_id = %s GROUP BY v.id, v.nombre, v.num_personas ORDER BY v.fecha_creacion DESC",
                    (user_id,),
                )
                vacas = cur.fetchall()
                cur.execute("SELECT persona, monto, monto_pagado, tasa_interes, descripcion, fecha, fecha_limite FROM prestamos WHERE user_id = %s AND pagado = FALSE ORDER BY fecha DESC", (user_id,))
                prestamos = cur.fetchall()

        wb = openpyxl.Workbook()
        hf = Font(bold=True, color="FFFFFF")
        fills = {
            "green":  PatternFill("solid", fgColor="1E7E34"),
            "blue":   PatternFill("solid", fgColor="0056B3"),
            "orange": PatternFill("solid", fgColor="D97000"),
            "purple": PatternFill("solid", fgColor="6F42C1"),
            "teal":   PatternFill("solid", fgColor="0097A7"),
        }
        center = Alignment(horizontal="center")

        def hdr(ws, fila, cols, color):
            for c in range(1, cols + 1):
                cell = ws.cell(row=fila, column=c)
                cell.font = hf
                cell.fill = fills[color]
                cell.alignment = center

        def autofit(ws):
            for col in ws.columns:
                w = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = w + 4

        # Hoja 1: Gastos
        ws1 = wb.active
        ws1.title = f"Gastos {mes}"
        ws1.append(["Categoría", "Monto", "Fecha"])
        hdr(ws1, 1, 3, "green")
        total_gastos = 0
        for g in gastos:
            ws1.append([g[0], float(g[1]), str(g[2])])
            total_gastos += float(g[1])
        ws1.append([])
        ws1.append(["TOTAL", total_gastos, ""])
        ws1.cell(ws1.max_row, 1).font = Font(bold=True)
        ws1.cell(ws1.max_row, 2).font = Font(bold=True)
        autofit(ws1)

        # Hoja 2: Ingresos
        ws2 = wb.create_sheet(f"Ingresos {mes}")
        ws2.append(["Descripción", "Monto", "Fecha"])
        hdr(ws2, 1, 3, "blue")
        total_ingresos = 0
        for i in ingresos:
            ws2.append([i[0], float(i[1]), str(i[2])])
            total_ingresos += float(i[1])
        ws2.append([])
        ws2.append(["TOTAL", total_ingresos, ""])
        ws2.cell(ws2.max_row, 1).font = Font(bold=True)
        ws2.cell(ws2.max_row, 2).font = Font(bold=True)
        autofit(ws2)

        # Hoja 3: Deudas y Vacas
        ws3 = wb.create_sheet("Deudas y Vacas")
        ws3.append(["DEUDAS PENDIENTES", "", "", "", ""])
        ws3.cell(1, 1).font = Font(bold=True, size=12)
        ws3.append(["Tipo", "Persona", "Monto", "Descripción", "Fecha"])
        hdr(ws3, 2, 5, "orange")
        for d in deudas:
            ws3.append(["Me deben" if d[0] == "prestado" else "Debo", d[1], float(d[2]), d[3], str(d[4])])
        ws3.append([])
        ws3.append(["VACAS GRUPALES", "", ""])
        ws3.cell(ws3.max_row, 1).font = Font(bold=True, size=12)
        ws3.append(["Nombre", "Personas", "Total"])
        hdr(ws3, ws3.max_row, 3, "purple")
        for v in vacas:
            ws3.append([v[0], v[1], float(v[2])])
        autofit(ws3)

        # Hoja 4: Préstamos
        ws_p = wb.create_sheet("Préstamos")
        ws_p.append(["Persona", "Monto", "Abonado", "Pendiente", "Interés %", "Descripción", "Fecha", "Vence"])
        hdr(ws_p, 1, 8, "teal")
        total_prestado = 0
        for p in prestamos:
            persona, monto, pagado, tasa, desc, fecha, limite = p
            monto = float(monto); pagado = float(pagado)
            pendiente = monto - pagado
            total_prestado += pendiente
            ws_p.append([persona, monto, pagado, pendiente, float(tasa), desc, str(fecha), str(limite) if limite else "Sin fecha"])
        ws_p.append([])
        ws_p.append(["TOTAL PENDIENTE", "", "", total_prestado, "", "", "", ""])
        ws_p.cell(ws_p.max_row, 1).font = Font(bold=True)
        ws_p.cell(ws_p.max_row, 4).font = Font(bold=True)
        autofit(ws_p)

        # Hoja 5: Resumen
        ws4 = wb.create_sheet("Resumen")
        balance = total_ingresos - total_gastos
        estado = "✅ Positivo" if balance >= 0 else "🔴 Negativo"
        ws4.append(["RESUMEN FINANCIERO", mes])
        ws4.cell(1, 1).font = Font(bold=True, size=14)
        ws4.append([])
        ws4.append(["Concepto", "Monto"])
        hdr(ws4, 3, 2, "blue")
        ws4.append(["Total Ingresos", total_ingresos])
        ws4.append(["Total Gastos", total_gastos])
        ws4.append(["Balance", balance])
        ws4.append(["Estado", estado])
        ws4.append([])
        ws4.append(["Deudas pendientes", len(deudas)])
        ws4.append(["Vacas activas", len(vacas)])
        ws4.append(["Total prestado pendiente", total_prestado])
        ws4.cell(6, 1).font = Font(bold=True)
        ws4.cell(6, 2).font = Font(bold=True, color="1E7E34" if balance >= 0 else "DC3545")
        autofit(ws4)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return {
            "output": f"📊 Reporte Excel de {mes} listo.",
            "excel_buffer": buffer.getvalue(),
            "excel_nombre": f"finanzas_{mes}.xlsx"
        }
    except Exception as e:
        print(f"[Error generar_excel] {e}")
        return {"output": "❌ Error generando el Excel. Intenta de nuevo."}


# ── Nodo 19: Registrar préstamo ────────────────────────────────────────────

def registrar_prestamo(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario prestó dinero a alguien y quiere hacer seguimiento formal con abonos.
        Extrae en JSON:
        {{
          "persona": "...",
          "monto": ...,
          "descripcion": "...",
          "tasa_interes": ... (% mensual, 0 si no menciona interés),
          "fecha_limite": "YYYY-MM-DD" (null si no menciona fecha límite)
        }}
        Si escribe "10k" interpreta como 10000.
        Fecha hoy: {date.today()}
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
        except Exception:
            return {"output": "No pude entender. Intenta: 'le presté 200000 a Carlos para el arriendo'"}

        persona      = params.get("persona", "")
        monto        = float(params.get("monto", 0))
        descripcion  = params.get("descripcion", "")
        tasa         = float(params.get("tasa_interes", 0))
        fecha_limite = params.get("fecha_limite")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO prestamos (user_id, persona, monto, descripcion, tasa_interes, fecha_limite) VALUES (%s, %s, %s, %s, %s, %s)",
                    (user_id, persona, monto, descripcion, tasa, fecha_limite),
                )

        interes_msg = f" con {tasa}% de interés mensual" if tasa > 0 else " sin interés"
        limite_msg  = f"\n📅 Fecha límite: {fecha_limite}" if fecha_limite else ""
        return {"output": (
            f"✅ Préstamo registrado\n"
            f"👤 Prestado a: *{persona}*\n"
            f"💵 Monto: ${int(monto):,}{interes_msg}\n"
            f"📝 {descripcion}{limite_msg}"
        )}
    except Exception as e:
        print(f"[Error registrar_prestamo] {e}")
        return {"output": "❌ Error registrando el préstamo."}


# ── Nodo 20: Ver préstamos ─────────────────────────────────────────────────

def ver_prestamos(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT persona, monto, monto_pagado, tasa_interes, descripcion, fecha, fecha_limite FROM prestamos WHERE user_id = %s AND pagado = FALSE ORDER BY fecha DESC",
                    (user_id,),
                )
                prestamos = cur.fetchall()

        if not prestamos:
            return {"output": "No tienes préstamos pendientes por cobrar 🎉"}

        respuesta = "💰 *Préstamos pendientes por cobrar*\n\n"
        total = 0

        for p in prestamos:
            persona, monto, pagado, tasa, desc, fecha, limite = p
            monto   = float(monto)
            pagado  = float(pagado)
            pendiente = monto - pagado

            if tasa and float(tasa) > 0:
                meses = max(1, (date.today() - fecha).days // 30)
                interes = monto * (float(tasa) / 100) * meses
                total_con_interes = pendiente + interes
                interes_str = f" (+${int(interes):,} interés acum.)"
            else:
                total_con_interes = pendiente
                interes_str = ""

            total += total_con_interes
            limite_str = f" | vence {limite}" if limite else ""
            abono_str  = f" | abonado: ${int(pagado):,}" if pagado > 0 else ""

            respuesta += (
                f"👤 *{persona}*: ${int(pendiente):,}{interes_str}{abono_str}{limite_str}\n"
                f"   📝 {desc}\n\n"
            )

        respuesta += f"💵 *Total por cobrar: ${int(total):,}*"
        return {"output": respuesta}
    except Exception as e:
        print(f"[Error ver_prestamos] {e}")
        return {"output": "❌ Error consultando los préstamos."}


# ── Nodo 21: Abonar préstamo ───────────────────────────────────────────────

def abonar_prestamo(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere registrar un abono a un préstamo que hizo.
        Extrae en JSON: {{"persona": "...", "monto_abono": ...}}
        Si escribe "10k" interpreta como 10000.
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            persona = params.get("persona", "")
            abono   = float(params.get("monto_abono", 0))
        except Exception:
            return {"output": "No pude entender. Intenta: 'Carlos me abonó 50000'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, monto, monto_pagado FROM prestamos WHERE user_id = %s AND LOWER(persona) LIKE LOWER(%s) AND pagado = FALSE ORDER BY id DESC LIMIT 1",
                    (user_id, f"%{persona}%"),
                )
                row = cur.fetchone()
                if not row:
                    return {"output": f"No encontré préstamo pendiente con '{persona}'."}

                prestamo_id  = row[0]
                monto_total  = float(row[1])
                ya_pagado    = float(row[2])
                nuevo_pagado = ya_pagado + abono
                pendiente    = monto_total - nuevo_pagado

                if nuevo_pagado >= monto_total:
                    cur.execute("UPDATE prestamos SET monto_pagado = %s, pagado = TRUE WHERE id = %s", (nuevo_pagado, prestamo_id))
                    return {"output": f"✅ *{persona}* pagó el préstamo completo 🎉\n💵 Total recibido: ${int(nuevo_pagado):,}"}
                else:
                    cur.execute("UPDATE prestamos SET monto_pagado = %s WHERE id = %s", (nuevo_pagado, prestamo_id))

        return {"output": (
            f"✅ Abono registrado de *{persona}*\n"
            f"💵 Abono: ${int(abono):,}\n"
            f"✔️ Total pagado: ${int(nuevo_pagado):,}\n"
            f"⏳ Pendiente: ${int(pendiente):,}"
        )}
    except Exception as e:
        print(f"[Error abonar_prestamo] {e}")
        return {"output": "❌ Error registrando el abono."}


# ── Nodo 22: Cerrar préstamo ───────────────────────────────────────────────

def cerrar_prestamo(state: dict) -> dict:
    try:
        user_id = _validar_user_id(state)
        texto = state["input"]

        prompt = f"""
        El usuario quiere marcar un préstamo como completamente pagado.
        Extrae en JSON: {{"persona": "..."}}
        Texto: "{texto}"
        """
        resp = client.chat.completions.create(
            model=MODELO_LLM,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
        )
        try:
            params = json.loads(resp.choices[0].message.content)
            persona = params.get("persona", "")
        except Exception:
            return {"output": "No pude entender. Intenta: 'Carlos pagó todo'"}

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE prestamos SET pagado = TRUE WHERE user_id = %s AND LOWER(persona) LIKE LOWER(%s) AND pagado = FALSE",
                    (user_id, f"%{persona}%"),
                )
                filas = cur.rowcount

        if filas:
            return {"output": f"✅ Préstamo con *{persona}* cerrado como pagado completo 🎉"}
        return {"output": f"No encontré préstamo pendiente con '{persona}'."}
    except Exception as e:
        print(f"[Error cerrar_prestamo] {e}")
        return {"output": "❌ Error cerrando el préstamo."}


# ── Nodo 23: Ayuda ────────────────────────────────────────────────────────

def ayuda(state: dict) -> dict:
    return {"output": (
        "🤖 *¿Qué puedo hacer por ti?*\n\n"
        "💸 *Gastos:* 'gasté 10k en comida' o 'gasté 10k en comida y 20k en netflix'\n"
        "📊 *Reportes:* 'reporte de hoy', 'reporte del mes', 'gastos por categoría'\n"
        "💰 *Presupuesto:* 'mi presupuesto es 500000', 'cómo voy con el presupuesto'\n"
        "💵 *Ingresos:* 'recibí salario 2000000', 'mis ingresos del mes'\n"
        "📈 *Balance:* 'cómo voy este mes', 'balance'\n"
        "🐄 *Vacas:* 'crear vaca Salida', 'agregar a vaca Salida: comida 50000, taxi 20000'\n"
        "🤝 *Deudas:* 'le debo 50000 a Juan', 'Juan me pagó', 'mis deudas'\n"
        "💳 *Préstamos:* 'le presté 200000 a Carlos', 'Carlos me abonó 50000'\n"
        "📋 *Excel:* 'generar excel', 'reporte de abril'\n"
    )}